"""
data_utils.py
=============
Data extraction and preprocessing utilities for the AMCS Two-Stage QUBO formulation.

Loads all required inputs from Scheduling_Input_2017.xlsx for use in the QUBO builder.

Sheet mapping (README + observed structure):
  C_INITIAL       – Initial C-check state per aircraft (FH/FC/DY elapsed, intervals, tolerances)
  A_INITIAL       – Initial A-check state per aircraft
  DFH             – Mean daily FH per aircraft × month
  DFC             – Mean daily FC per aircraft × month
  FH_STD          – Std-dev of daily FH per aircraft × month
  FH_MAX          – Max daily FH per aircraft × month (simulation clamp)
  FH_MIN          – Min daily FH per aircraft × month (simulation clamp)
  C_CHECK_CODE    – Upcoming C-check code sequence per aircraft (up to 5 checks)
  C_ELAPSED_TIME  – Expected TAT (days) for each check in the code sequence
  STOCHASTIC      – Empirical duration samples for every C-check code
  ADDITIONAL      – Scalar planning parameters (horizon, capacities, costs, ∆_min)
  C_PEAK          – Annual peak season blackout windows for C-checks (Jun–Sep)
  C_NOT_ALLOWED   – Holiday-period blackout windows for C-checks
  MORE_C_SLOTS    – Date-range overrides: total C-check hangar capacity for that window
  PUBLIC_HOLIDAYS – Individual public holiday dates (used in A-check blackout)
  A_NOT_ALLOWED   – Per-date A-check blackout (weekends, public holidays)
  MORE_A_SLOTS    – Per-date overrides: total A-check hangar capacity for that date
  RESERVED        – Pre-reserved maintenance slots (currently empty in this dataset)
"""

from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple

import numpy as np
import openpyxl

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
          "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]


def _to_date(val) -> Optional[datetime.date]:
    """Coerce openpyxl cell values (datetime, date, or None) to date."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    raise TypeError(f"Cannot convert {type(val)} to date: {val!r}")


def date_to_day(d: datetime.date, begin_day: datetime.date) -> int:
    """Return 0-based day index relative to the planning horizon start."""
    return (d - begin_day).days


def day_to_date(t: int, begin_day: datetime.date) -> datetime.date:
    """Inverse of date_to_day."""
    return begin_day + datetime.timedelta(days=t)


def _iter_data_rows(ws):
    """Yield all non-header, non-empty rows from a worksheet (skip first row)."""
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if all(v is None for v in row):
            continue
        yield row


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class AircraftCState:
    """Initial C-check state for a single aircraft at Day 0 (25 Sep 2017)."""
    fleet: str
    tail: str
    # Current check position in the sequence
    c_sn: float          # C-check serial number at Day 0 (e.g. 11.1 = 11th check, type 1)
    c_max: int           # Total check slots available for this aircraft in the horizon
    # Accumulated utilization since last C-check
    dy_c: float          # Elapsed calendar days since last C-check
    fh_c: float          # Cumulative flight hours since last C-check
    fc_c: float          # Cumulative flight cycles since last C-check
    # Check interval limits (from airworthiness directives)
    ci_dy: float         # Maximum interval (days)
    ci_fh: float         # Maximum interval (FH)
    ci_fc: float         # Maximum interval (FC)
    # Tolerance (how much we can exceed the nominal interval)
    tol_dy: float
    tol_fh: float
    tol_fc: float
    # Tolerance already consumed by the previous check
    tolu_dy: float
    tolu_fh: float
    tolu_fc: float
    # Pre-committed check window (-1 if no committed check yet)
    c_start: int         # Day index of committed C-check start (or -1)
    c_end: int           # Day index of committed C-check end   (or -1)

    @property
    def eff_limit_dy(self) -> float:
        """Effective DY limit = interval + remaining tolerance."""
        return self.ci_dy + self.tol_dy - self.tolu_dy

    @property
    def eff_limit_fh(self) -> float:
        """Effective FH limit = interval + remaining tolerance."""
        return self.ci_fh + self.tol_fh - self.tolu_fh

    @property
    def eff_limit_fc(self) -> float:
        """Effective FC limit = interval + remaining tolerance."""
        return self.ci_fc + self.tol_fc - self.tolu_fc


@dataclass
class AircraftAState:
    """Initial A-check state for a single aircraft at Day 0."""
    fleet: str
    tail: str
    a_sn: float          # A-check serial number at Day 0
    a_max: int           # Total A-check slots in the horizon
    dy_a: float          # Elapsed days since last A-check
    fh_a: float          # Cumulative FH since last A-check
    fc_a: float          # Cumulative FC since last A-check
    ci_dy: float
    ci_fh: float
    ci_fc: float
    tol_dy: float
    tol_fh: float
    tol_fc: float
    tolu_dy: float
    tolu_fh: float
    tolu_fc: float

    @property
    def eff_limit_dy(self) -> float:
        return self.ci_dy + self.tol_dy - self.tolu_dy

    @property
    def eff_limit_fh(self) -> float:
        return self.ci_fh + self.tol_fh - self.tolu_fh

    @property
    def eff_limit_fc(self) -> float:
        return self.ci_fc + self.tol_fc - self.tolu_fc


@dataclass
class PlanningParameters:
    """Scalar planning parameters from the ADDITIONAL sheet."""
    begin_year: int
    total_years: int
    begin_day: datetime.date      # Day 0 of the horizon
    m_cost: float                 # Penalty for missing a check (grounding cost)
    c_cost: float                 # Cost for opening an extra maintenance slot
    max_c_check: int              # Nominal C-check hangar capacity M_C
    max_a_check: int              # Nominal A-check hangar capacity M_A
    start_day_interval: int       # Minimum days between C-check start dates (∆_min)

    @property
    def horizon_days(self) -> int:
        """Total planning horizon in days (3 years ≈ 1096)."""
        # Count exact days for the 3-year span
        end = datetime.date(self.begin_year + self.total_years, 9, 25)
        return (end - self.begin_day).days


@dataclass
class StochasticDuration:
    """
    Empirical C-check duration distribution for a specific check code.

    The STOCHASTIC sheet stores one row per check code with a variable-length
    list of observed TAT (turnaround time) values, None-padded to a fixed width.
    """
    check_code: float            # e.g. 1.1, 2.1, ..., 12.2
    samples: List[int]           # Historical TAT observations (days)

    @property
    def mean(self) -> float:
        return float(np.mean(self.samples))

    @property
    def std(self) -> float:
        return float(np.std(self.samples, ddof=1)) if len(self.samples) > 1 else 0.0

    @property
    def min_val(self) -> int:
        return min(self.samples)

    @property
    def max_val(self) -> int:
        return max(self.samples)


@dataclass
class CapacityCalendar:
    """
    Per-day C-check and A-check hangar capacity over the full horizon.

    Accounts for:
      - Base capacity (M_C, M_A from ADDITIONAL)
      - Capacity overrides from MORE_C_SLOTS and MORE_A_SLOTS
      - Blackout days where capacity is 0 (C_PEAK + C_NOT_ALLOWED for C-checks;
        A_NOT_ALLOWED for A-checks), unless overridden by MORE_*_SLOTS.

    Blackout sets (BC, BA) contain day indices where effective capacity == 0.
    """
    # Indexed by day t (0-based from begin_day), length = horizon_days
    c_capacity: np.ndarray   # dtype=int, shape=(T,); 0 = blackout
    a_capacity: np.ndarray   # dtype=int, shape=(T,); 0 = blackout

    @property
    def blackout_c(self) -> Set[int]:
        """Set of day indices that are blacked out for C-checks."""
        return set(int(t) for t in np.where(self.c_capacity == 0)[0])

    @property
    def blackout_a(self) -> Set[int]:
        """Set of day indices that are blacked out for A-checks."""
        return set(int(t) for t in np.where(self.a_capacity == 0)[0])


@dataclass
class AMCSData:
    """
    Complete dataset for the Two-Stage AMCS QUBO problem.

    All per-aircraft dicts are keyed by tail string ('Aircraft-1', etc.).
    Monthly arrays have shape (12,) with index 0 = January.
    """
    # Aircraft ordered list
    aircraft: List[str]

    # --- Initial states ---
    c_initial: Dict[str, AircraftCState]   # tail -> C-check state at Day 0
    a_initial: Dict[str, AircraftAState]   # tail -> A-check state at Day 0

    # --- Utilization parameters (per aircraft, per month) ---
    dfh: Dict[str, np.ndarray]    # Mean daily FH:  µ^FH_i(m)
    dfc: Dict[str, np.ndarray]    # Mean daily FC:  µ^FC_i(m)
    fh_std: Dict[str, np.ndarray] # Std-dev daily FH: σ^FH_i(m)
    fh_max: Dict[str, np.ndarray] # Max daily FH (simulation clamp)
    fh_min: Dict[str, np.ndarray] # Min daily FH (simulation clamp)

    # --- C-check sequences ---
    # Upcoming check codes per aircraft (k=0 is the NEXT check after Day 0)
    c_check_codes: Dict[str, List[float]]
    # Expected TAT (days) for each check in the sequence; -1 = unknown
    c_elapsed_times: Dict[str, List[int]]

    # --- Duration distributions (for simulation sampling) ---
    stochastic: Dict[float, StochasticDuration]  # check_code -> distribution

    # --- Planning parameters ---
    params: PlanningParameters

    # --- Capacity calendar (C and A) ---
    capacity: CapacityCalendar

    def get_c_duration(self, tail: str, k: int) -> int:
        """
        Return expected C-check duration for aircraft `tail`, check index k
        (0-based within the upcoming sequence).

        Priority:
          1. C_ELAPSED_TIME value if > 0.
          2. STOCHASTIC mean for the matching check code.
          3. Fallback: 13 days (conservative middle estimate).
        """
        tat = self.c_elapsed_times[tail][k]
        if tat > 0:
            return tat
        code = self.c_check_codes[tail][k]
        if code in self.stochastic:
            return max(1, round(self.stochastic[code].mean))
        return 13  # fallback

    def get_stochastic_duration(self, tail: str, k: int) -> StochasticDuration:
        """
        Return StochasticDuration for aircraft `tail`, check index k.
        Falls back to a degenerate distribution using the TAT value.
        """
        code = self.c_check_codes[tail][k]
        if code in self.stochastic:
            return self.stochastic[code]
        tat = self.c_elapsed_times[tail][k]
        dur = tat if tat > 0 else 13
        return StochasticDuration(check_code=code, samples=[dur])


# ---------------------------------------------------------------------------
# Sheet loaders (private)
# ---------------------------------------------------------------------------

def _load_c_initial(ws, begin_day: datetime.date) -> Tuple[List[str], Dict[str, AircraftCState]]:
    """Load C_INITIAL sheet → ordered aircraft list + dict of AircraftCState."""
    aircraft: List[str] = []
    states: Dict[str, AircraftCState] = {}

    for row in _iter_data_rows(ws):
        (fleet, tail, c_sn, c_max,
         dy_c, fh_c, fc_c,
         ci_dy, ci_fh, ci_fc,
         tol_dy, tol_fh, tol_fc,
         tolu_dy, tolu_fh, tolu_fc,
         c_start_raw, c_end_raw) = row

        # c_start/c_end are stored as day-index integers (-1 = not committed)
        # They may also be stored as dates in some versions – handle both.
        def _parse_day(val) -> int:
            if val is None or val == -1:
                return -1
            if isinstance(val, (int, float)):
                v = int(val)
                return v  # already a day index or -1
            d = _to_date(val)
            return date_to_day(d, begin_day) if d is not None else -1

        state = AircraftCState(
            fleet=str(fleet),
            tail=str(tail),
            c_sn=float(c_sn),
            c_max=int(c_max),
            dy_c=float(dy_c),
            fh_c=float(fh_c),
            fc_c=float(fc_c),
            ci_dy=float(ci_dy),
            ci_fh=float(ci_fh),
            ci_fc=float(ci_fc),
            tol_dy=float(tol_dy),
            tol_fh=float(tol_fh),
            tol_fc=float(tol_fc),
            tolu_dy=float(tolu_dy),
            tolu_fh=float(tolu_fh),
            tolu_fc=float(tolu_fc),
            c_start=_parse_day(c_start_raw),
            c_end=_parse_day(c_end_raw),
        )
        aircraft.append(tail)
        states[tail] = state

    return aircraft, states


def _load_a_initial(ws) -> Dict[str, AircraftAState]:
    """Load A_INITIAL sheet → dict of AircraftAState."""
    states: Dict[str, AircraftAState] = {}

    for row in _iter_data_rows(ws):
        (fleet, tail, a_sn, a_max,
         dy_a, fh_a, fc_a,
         ci_dy, ci_fh, ci_fc,
         tol_dy, tol_fh, tol_fc,
         tolu_dy, tolu_fh, tolu_fc) = row

        state = AircraftAState(
            fleet=str(fleet),
            tail=str(tail),
            a_sn=float(a_sn),
            a_max=int(a_max),
            dy_a=float(dy_a),
            fh_a=float(fh_a),
            fc_a=float(fc_a),
            ci_dy=float(ci_dy),
            ci_fh=float(ci_fh),
            ci_fc=float(ci_fc),
            tol_dy=float(tol_dy),
            tol_fh=float(tol_fh),
            tol_fc=float(tol_fc),
            tolu_dy=float(tolu_dy),
            tolu_fh=float(tolu_fh),
            tolu_fc=float(tolu_fc),
        )
        states[tail] = state

    return states


def _load_monthly_data(ws) -> Dict[str, np.ndarray]:
    """
    Load a sheet with columns [FLEET, A/C TAIL, JAN, FEB, ..., DEC].
    Returns dict: tail -> float array of shape (12,), index 0 = January.
    """
    data: Dict[str, np.ndarray] = {}
    for row in _iter_data_rows(ws):
        tail = row[1]
        if tail is None:
            continue
        monthly = np.array([float(v) if v is not None else np.nan
                            for v in row[2:14]], dtype=float)
        data[str(tail)] = monthly
    return data


def _load_c_check_codes(ws) -> Dict[str, List[float]]:
    """
    Load C_CHECK_CODE sheet.
    Columns: FLEET, A/C TAIL, C CODE (×5 upcoming checks).
    Returns dict: tail -> list of check codes (float, e.g. 11.1, 1.2).
    Only non-None codes are included.
    """
    codes: Dict[str, List[float]] = {}
    for row in _iter_data_rows(ws):
        tail = str(row[1])
        check_codes = [float(v) for v in row[2:] if v is not None]
        codes[tail] = check_codes
    return codes


def _load_c_elapsed_times(ws) -> Dict[str, List[int]]:
    """
    Load C_ELAPSED_TIME sheet.
    Columns: FLEET, A/C TAIL, TAT (×5).
    Returns dict: tail -> list of TAT values.
    -1 means unknown/not applicable for that check slot.
    """
    times: Dict[str, List[int]] = {}
    for row in _iter_data_rows(ws):
        tail = str(row[1])
        tats = [int(v) if v is not None else -1 for v in row[2:]]
        # Trim trailing -1s that correspond to non-existent future checks
        while tats and tats[-1] == -1:
            tats.pop()
        times[tail] = tats
    return times


def _load_stochastic(ws) -> Dict[float, StochasticDuration]:
    """
    Load STOCHASTIC sheet.
    Row format: (fleet_label, 'C-Check', check_code, sample1, sample2, ..., NoneNone...)
    Returns dict: check_code (float) -> StochasticDuration.
    """
    distributions: Dict[float, StochasticDuration] = {}
    for row in ws.iter_rows(values_only=True):
        if row[2] is None:
            continue
        check_code = float(row[2])
        samples = [int(v) for v in row[3:] if v is not None]
        if not samples:
            continue
        distributions[check_code] = StochasticDuration(
            check_code=check_code,
            samples=samples,
        )
    return distributions


def _load_planning_params(ws) -> PlanningParameters:
    """Load ADDITIONAL sheet → PlanningParameters."""
    data: Dict[str, object] = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None:
            data[str(row[0])] = row[1]

    begin_day_raw = data["BEGIN DAY"]
    begin_day = _to_date(begin_day_raw)

    return PlanningParameters(
        begin_year=int(data["BEGIN YEAR"]),
        total_years=int(data["TOTAL YEARS"]),
        begin_day=begin_day,
        m_cost=float(data["M COST"]),
        c_cost=float(data["C COST"]),
        max_c_check=int(data["MAX C CHECK"]),
        max_a_check=int(data["MAX A CHECK"]),
        start_day_interval=int(data["START DAY INTERVAL"]),
    )


def _build_capacity_calendar(
    c_peak_ws,
    c_not_allowed_ws,
    more_c_slots_ws,
    a_not_allowed_ws,
    more_a_slots_ws,
    params: PlanningParameters,
) -> CapacityCalendar:
    """
    Build per-day hangar capacity arrays for C-checks and A-checks.

    Logic:
      C-checks:
        1. Start with base capacity M_C on every day.
        2. Apply blackouts from C_PEAK (Jun–Sep each year) → set to 0.
        3. Apply blackouts from C_NOT_ALLOWED (holiday windows) → set to 0.
        4. Apply MORE_C_SLOTS overrides (TOTAL SLOTS for date ranges) → override.
           These can restore capacity on blacked-out days OR reduce it below M_C.

      A-checks:
        1. Start with base capacity M_A on every day.
        2. Apply blackouts from A_NOT_ALLOWED (specific dates) → set to 0.
        3. Apply MORE_A_SLOTS overrides (TOTAL SLOTS for specific dates) → override.
    """
    T = params.horizon_days
    begin_day = params.begin_day

    c_cap = np.full(T, params.max_c_check, dtype=int)
    a_cap = np.full(T, params.max_a_check, dtype=int)

    # --- C-check blackouts: C_PEAK ---
    for row in _iter_data_rows(c_peak_ws):
        _year, peak_begin, peak_end = row[0], _to_date(row[1]), _to_date(row[2])
        if peak_begin is None or peak_end is None:
            continue
        t0 = max(0, date_to_day(peak_begin, begin_day))
        t1 = min(T - 1, date_to_day(peak_end, begin_day))
        if t0 <= t1:
            c_cap[t0:t1 + 1] = 0

    # --- C-check blackouts: C_NOT_ALLOWED ---
    for row in _iter_data_rows(c_not_allowed_ws):
        start, end = _to_date(row[0]), _to_date(row[1])
        if start is None or end is None:
            continue
        t0 = max(0, date_to_day(start, begin_day))
        t1 = min(T - 1, date_to_day(end, begin_day))
        if t0 <= t1:
            c_cap[t0:t1 + 1] = 0

    # --- C-check capacity overrides: MORE_C_SLOTS ---
    # These take precedence over blackouts (applied last).
    for row in _iter_data_rows(more_c_slots_ws):
        start, end, total_slots = _to_date(row[0]), _to_date(row[1]), row[2]
        if start is None or end is None or total_slots is None:
            continue
        t0 = max(0, date_to_day(start, begin_day))
        t1 = min(T - 1, date_to_day(end, begin_day))
        if t0 <= t1:
            c_cap[t0:t1 + 1] = int(total_slots)

    # --- A-check blackouts: A_NOT_ALLOWED ---
    for row in _iter_data_rows(a_not_allowed_ws):
        d = _to_date(row[0])
        if d is None:
            continue
        t = date_to_day(d, begin_day)
        if 0 <= t < T:
            a_cap[t] = 0

    # --- A-check capacity overrides: MORE_A_SLOTS ---
    for row in _iter_data_rows(more_a_slots_ws):
        d, total_slots = _to_date(row[0]), row[1]
        if d is None or total_slots is None:
            continue
        t = date_to_day(d, begin_day)
        if 0 <= t < T:
            a_cap[t] = int(total_slots)

    return CapacityCalendar(c_capacity=c_cap, a_capacity=a_cap)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_amcs_data(filepath: str) -> AMCSData:
    """
    Load all data from the AMCS scheduling Excel workbook.

    Parameters
    ----------
    filepath : str
        Path to Scheduling_Input_2017.xlsx (or the 2019-2022 variant).

    Returns
    -------
    AMCSData
        Fully populated dataset ready for the QUBO builder.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # 1. Planning parameters (needed first to establish begin_day / horizon)
    params = _load_planning_params(wb["ADDITIONAL"])

    # 2. Aircraft initial states
    aircraft, c_initial = _load_c_initial(wb["C_INITIAL"], params.begin_day)
    a_initial = _load_a_initial(wb["A_INITIAL"])

    # 3. Utilization parameters (monthly, per aircraft)
    dfh    = _load_monthly_data(wb["DFH"])
    dfc    = _load_monthly_data(wb["DFC"])
    fh_std = _load_monthly_data(wb["FH_STD"])
    fh_max = _load_monthly_data(wb["FH_MAX"])
    fh_min = _load_monthly_data(wb["FH_MIN"])

    # 4. C-check code sequences and expected durations
    c_check_codes    = _load_c_check_codes(wb["C_CHECK_CODE"])
    c_elapsed_times  = _load_c_elapsed_times(wb["C_ELAPSED_TIME"])

    # Align elapsed times list length to match check codes list per aircraft
    for tail in aircraft:
        codes = c_check_codes.get(tail, [])
        tats  = c_elapsed_times.get(tail, [])
        # Pad with -1 if TAT list is shorter than code list
        while len(tats) < len(codes):
            tats.append(-1)
        c_elapsed_times[tail] = tats[:len(codes)]

    # 5. Stochastic duration distributions
    stochastic = _load_stochastic(wb["STOCHASTIC"])

    # 6. Capacity calendar (blackouts + overrides)
    capacity = _build_capacity_calendar(
        c_peak_ws=wb["C_PEAK"],
        c_not_allowed_ws=wb["C_NOT_ALLOWED"],
        more_c_slots_ws=wb["MORE_C_SLOTS"],
        a_not_allowed_ws=wb["A_NOT_ALLOWED"],
        more_a_slots_ws=wb["MORE_A_SLOTS"],
        params=params,
    )

    wb.close()

    return AMCSData(
        aircraft=aircraft,
        c_initial=c_initial,
        a_initial=a_initial,
        dfh=dfh,
        dfc=dfc,
        fh_std=fh_std,
        fh_max=fh_max,
        fh_min=fh_min,
        c_check_codes=c_check_codes,
        c_elapsed_times=c_elapsed_times,
        stochastic=stochastic,
        params=params,
        capacity=capacity,
    )


# ---------------------------------------------------------------------------
# Diagnostic / sanity-check
# ---------------------------------------------------------------------------

def print_summary(data: AMCSData) -> None:
    """Print a concise summary of the loaded dataset for quick validation."""
    p = data.params
    T = p.horizon_days

    print("=" * 60)
    print("AMCS Dataset Summary")
    print("=" * 60)
    end_date = p.begin_day + datetime.timedelta(days=T - 1)
    print(f"  Planning horizon : {p.begin_day} to {end_date}  ({T} days)")
    print(f"  Aircraft         : {len(data.aircraft)}")
    print(f"  M_C (C-hangar)   : {p.max_c_check}   d_min = {p.start_day_interval} days")
    print(f"  M_A (A-hangar)   : {p.max_a_check}")
    print()

    # C-check state sample
    print("  C-check initial states (first 3 aircraft):")
    for tail in data.aircraft[:3]:
        cs = data.c_initial[tail]
        print(f"    {tail} ({cs.fleet}): "
              f"DY={cs.dy_c:.0f}/{cs.eff_limit_dy:.0f}  "
              f"FH={cs.fh_c:.0f}/{cs.eff_limit_fh:.0f}  "
              f"FC={cs.fc_c:.0f}/{cs.eff_limit_fc:.0f}  "
              f"next_codes={data.c_check_codes[tail][:2]}")
    print()

    # A-check state sample
    print("  A-check initial states (first 3 aircraft):")
    for tail in data.aircraft[:3]:
        as_ = data.a_initial[tail]
        print(f"    {tail}: "
              f"DY={as_.dy_a:.0f}/{as_.eff_limit_dy:.0f}  "
              f"FH={as_.fh_a:.0f}/{as_.eff_limit_fh:.0f}  "
              f"FC={as_.fc_a:.0f}/{as_.eff_limit_fc:.0f}")
    print()

    # Utilization sample
    tail0 = data.aircraft[0]
    print(f"  DFH for {tail0}: {data.dfh[tail0].round(2)}")
    print(f"  DFC for {tail0}: {data.dfc[tail0].round(2)}")
    print()

    # Stochastic durations
    print(f"  Stochastic distributions loaded: {len(data.stochastic)} check codes")
    for code, dist in sorted(data.stochastic.items()):
        print(f"    {code:4.1f}: n={len(dist.samples):2d}  "
              f"mean={dist.mean:5.1f}  std={dist.std:4.1f}  "
              f"range=[{dist.min_val},{dist.max_val}]")
    print()

    # Capacity calendar
    c_blk = len(data.capacity.blackout_c)
    a_blk = len(data.capacity.blackout_a)
    print(f"  C-check blackout days : {c_blk} / {T} ({100*c_blk/T:.1f}%)")
    print(f"  A-check blackout days : {a_blk} / {T} ({100*a_blk/T:.1f}%)")

    # Extra capacity days (> base M_C or M_A)
    c_extra = int(np.sum(data.capacity.c_capacity > p.max_c_check))
    a_extra = int(np.sum(data.capacity.a_capacity > p.max_a_check))
    print(f"  C-check extra-capacity days: {c_extra}")
    print(f"  A-check extra-capacity days: {a_extra}")
    print("=" * 60)


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "Scheduling_Input_2017.xlsx"
    data = load_amcs_data(path)
    print_summary(data)
